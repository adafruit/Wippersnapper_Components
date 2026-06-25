#!/usr/bin/env python3
"""Generate HIL testing spreadsheet for all I2C Wippersnapper components.

Creates an Excel workbook with:
  Sheet 1 - Component Matrix: all components, addresses, vendors, conflicts
  Sheet 2 - HIL Mux Layout: conflict-free channel assignments for testing
  Sheet 3 - Address Conflicts: per-address overlap summary
  Sheet 4 - Test Fixtures: components grouped by measured phenomena

Hardware config:
  - 8-ch TCA9548A mux @ 0x77  (all 3 addr pads bridged)
  - 4-ch TCA9544A mux @ 0x71  (A0 bridged)
  - Reserved addresses on every channel: {0x77, 0x71}
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
import time
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Hardware config ──────────────────────────────────────────────────────────
# Full mux inventory. In practice the HIL rig drives a SINGLE 8-channel mux —
# two muxes can't realistically be used at once — so the second 4-channel mux
# (TCA9544A @ 0x71) is opt-in via --dual-mux and OFF by default. When it is off,
# 0x71 is treated as a normal, usable component address.
ALL_MUXES = [
    {"type": "TCA9548A", "address": 0x77, "channels": 8, "label": "8ch"},
    {"type": "TCA9544A", "address": 0x71, "channels": 4, "label": "4ch"},
]
USE_SECOND_MUX = False  # default: single mux (TCA9548A @ 0x77) only

# Active mux globals — (re)computed by configure_muxes() from USE_SECOND_MUX.
MUXES = []
MUX_RESERVED = set()
TOTAL_MUX_CHANNELS = 0


def configure_muxes(use_second_mux):
    """Select the active muxes. With use_second_mux=False (default) only the
    8-channel TCA9548A @ 0x77 is used and 0x71 becomes a normal usable address."""
    global MUXES, MUX_RESERVED, TOTAL_MUX_CHANNELS, USE_SECOND_MUX
    USE_SECOND_MUX = use_second_mux
    MUXES = list(ALL_MUXES) if use_second_mux else ALL_MUXES[:1]
    MUX_RESERVED = {m["address"] for m in MUXES}
    TOTAL_MUX_CHANNELS = sum(m["channels"] for m in MUXES)


configure_muxes(USE_SECOND_MUX)


# ── Colours ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CONFLICT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
UNIQUE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
UNPUBLISHED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
# Strong amber to make an unpublished ("NO") Published cell stand out from the
# faint row tint above.
UNPUBLISHED_CELL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
MUX_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
MUX2_HEADER_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
CHANNEL_FILLS_MUX1 = [
    PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # Ch0
    PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # Ch1
    PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # Ch2
    PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # Ch3
    PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # Ch4
    PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # Ch5
    PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # Ch6
    PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # Ch7
]
CHANNEL_FILLS_MUX2 = [
    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # Ch0
    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Ch1
    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # Ch2
    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Ch3
]
NON_DEFAULT_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # orange — assigned != default
NON_DEFAULT_FONT = Font(bold=True, color="833C0B")
NOMUX_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
NOMUX_HEADER_FILL = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


JUMPER_JSON_PATH = "i2c_address_jumper_info.json"


def load_jumper_info(base_dir):
    """Load cached address jumper info from JSON."""
    path = Path(base_dir) / JUMPER_JSON_PATH
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("components", {})
    return {}


def save_jumper_info(base_dir, jumper_db):
    """Write jumper info back to the JSON cache."""
    path = Path(base_dir) / JUMPER_JSON_PATH
    # Read existing file to preserve metadata
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {"metadata": {}, "components": {}}
    data["components"] = jumper_db
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")


def _find_claude_cli():
    """Locate the claude CLI across Windows, macOS, Linux, and WSL."""
    # Try plain 'claude' first (works on macOS/Linux, and Windows if in PATH)
    cmd = shutil.which("claude")
    if cmd:
        return cmd
    # On Windows, npm global installs create .cmd shims
    if sys.platform == "win32":
        for ext in (".cmd", ".exe", ".ps1"):
            cmd = shutil.which("claude" + ext)
            if cmd:
                return cmd
        # Also check common npm global locations
        for candidate in [
            Path(os.environ.get("APPDATA", "")) / "npm" / "claude.cmd",
            Path(os.environ.get("LOCALAPPDATA", "")) / "npm" / "claude.cmd",
        ]:
            if candidate.exists():
                return str(candidate)
    return None


def fetch_jumper_info_via_claude(component_name, doc_url, addresses):
    """Call 'claude -p' to research address jumper info for a component."""
    addrs_str = ", ".join(addresses)
    prompt = (
        f"For the Adafruit I2C breakout board '{component_name}' "
        f"(I2C addresses: {addrs_str}), "
        f"documentation: {doc_url or 'none'}\n\n"
        f"What is the I2C address jumper/solder pad configuration? "
        f"Return ONLY a JSON object (no markdown fencing) with these fields:\n"
        f'  "default_address": hex string like "0x48",\n'
        f'  "has_address_jumper": true/false,\n'
        f'  "address_jumper_info": concise description of how to change address,\n'
        f'  "addresses": object mapping hex address to description,\n'
        f'  "confidence": "high" or "low"\n'
    )

    claude_cmd = _find_claude_cli()
    if not claude_cmd:
        print("    'claude' CLI not found — skipping auto-fetch")
        return None

    try:
        # Use shell=True on Windows for .cmd shims; direct exec elsewhere
        use_shell = sys.platform == "win32" and claude_cmd.endswith(".cmd")
        result = subprocess.run(
            [claude_cmd, "-p", prompt, "--output-format", "text"],
            capture_output=True, text=True, timeout=120,
            shell=use_shell,
        )
        if result.returncode != 0:
            print(f"    claude exited with code {result.returncode}")
            if result.stderr:
                print(f"    stderr: {result.stderr[:200]}")
            return None

        # Parse JSON from output (strip any markdown fencing)
        output = result.stdout.strip()
        if output.startswith("```"):
            output = "\n".join(output.split("\n")[1:])
        if output.endswith("```"):
            output = "\n".join(output.split("\n")[:-1])
        output = output.strip()

        data = json.loads(output)
        return data
    except subprocess.TimeoutExpired:
        print("    claude timed out (120s)")
        return None
    except json.JSONDecodeError as e:
        print(f"    failed to parse claude output as JSON: {e}")
        print(f"    raw output: {output[:300]}")
        return None
    except Exception as e:
        print(f"    error calling claude: {e}")
        return None


def check_and_fetch_missing_jumper_info(base_dir, components, jumper_db):
    """Check for components missing jumper info, offer to fetch via claude -p."""
    missing = []
    for comp in components:
        name = comp["dir"]
        if name not in jumper_db:
            doc_url = comp.get("guide_url", "")
            addrs = [f"0x{a:02X}" for a in comp["all_addresses"]]
            missing.append((name, doc_url, addrs))

    if not missing:
        return jumper_db

    print(f"\n{'='*60}")
    print(f"WARNING: {len(missing)} component(s) missing address jumper info:")
    for name, doc_url, addrs in missing:
        print(f"  - {name} ({', '.join(addrs)})")
    print(f"{'='*60}")

    if not _find_claude_cli():
        print("'claude' CLI not found — cannot auto-fetch. Add entries manually to")
        print(f"  {JUMPER_JSON_PATH}")
        return jumper_db

    print(f"\nWill attempt to fetch info via 'claude -p' in 5 seconds...")
    print("Press Ctrl+C to skip.\n")
    try:
        for i in range(5, 0, -1):
            print(f"  {i}...", end=" ", flush=True)
            time.sleep(1)
        print()
    except KeyboardInterrupt:
        print("\n  Skipped.")
        return jumper_db

    updated = False
    for name, doc_url, addrs in missing:
        print(f"\n  Fetching jumper info for '{name}'...")
        data = fetch_jumper_info_via_claude(name, doc_url, addrs)
        if data:
            jumper_db[name] = {
                "default_address": data.get("default_address", addrs[0] if addrs else "?"),
                "has_address_jumper": data.get("has_address_jumper", False),
                "address_jumper_info": data.get("address_jumper_info", ""),
                "addresses": data.get("addresses", {}),
                "guide_url": doc_url or None,
                "confidence": data.get("confidence", "low"),
                "auto_fetched": True,
            }
            print(f"    OK: {data.get('address_jumper_info', '')[:80]}")
            updated = True
        else:
            print(f"    FAILED — add manually to {JUMPER_JSON_PATH}")

    if updated:
        save_jumper_info(base_dir, jumper_db)
        print(f"\n  Updated {JUMPER_JSON_PATH}")

    return jumper_db


def load_components(base_dir):
    """Load all I2C component definitions."""
    jumper_db = load_jumper_info(base_dir)
    components = []
    i2c_dir = Path(base_dir) / "components" / "i2c"
    for defn in sorted(i2c_dir.glob("*/definition.json")):
        with open(defn, "r", encoding="utf-8") as f:
            data = json.load(f)
        name = defn.parent.name
        addrs = [int(a, 16) for a in data.get("i2cAddresses", [])]  # preserve definition order (default first)
        usable = [a for a in addrs if a not in MUX_RESERVED]
        jinfo = jumper_db.get(name, {})
        components.append({
            "dir": name,
            "displayName": data.get("displayName", name),
            "vendor": data.get("vendorName", ""),
            "all_addresses": addrs,
            "usable_addresses": usable,
            "published": data.get("published", True),
            "sensors": [
                s["sensorType"] if isinstance(s, dict) else s
                for s in data.get("subcomponents", [])
            ],
            "has_jumper": jinfo.get("has_address_jumper", None),
            "jumper_info": jinfo.get("address_jumper_info", ""),
            "jumper_addrs": jinfo.get("addresses", {}),  # {hex_addr: description}
            "guide_url": jinfo.get("guide_url") or data.get("documentationURL", ""),
        })
    components.sort(key=lambda c: (c["all_addresses"][0] if c["all_addresses"] else 0xFF, c["dir"]))
    return components


def build_address_map(components):
    """Map address -> list of component dicts (using all_addresses)."""
    addr_map = defaultdict(list)
    for comp in components:
        for a in comp["all_addresses"]:
            addr_map[a].append(comp)
    return addr_map


def find_conflicts(components, addr_map):
    """For each component, find which other components share any address."""
    conflicts = {}
    for comp in components:
        peers = set()
        for a in comp["all_addresses"]:
            for other in addr_map[a]:
                if other["dir"] != comp["dir"]:
                    peers.add(other["dir"])
        conflicts[comp["dir"]] = sorted(peers)
    return conflicts


# ── Channel assignment with single-address picking ───────────────────────────

def assign_channels(components):
    """
    Assign each component to a channel AND pick ONE specific address.

    Strategy:
      1. Put components with truly unique addresses on the direct bus (ch 0).
         "Truly unique" = the component has a usable address that NO other
         component lists in its usable set.  These are free on the direct bus
         because they block an address nobody else needs.
      2. Everything else goes onto mux channels (1-12) via greedy colouring.
         Most-constrained-first (fewest usable addresses).

    Channel layout:
      0           = direct bus (no mux) — always visible
      1  .. 8     = 8ch TCA9548A @ 0x77, channels 0-7
      9  .. 12    = 4ch TCA9544A @ 0x71, channels 0-3

    Constraints:
      - Picked address must not be in MUX_RESERVED
      - No two components on the same channel share a picked address
      - No muxed component's picked address clashes with any direct-bus
        component's picked address (direct bus is always visible)

    Returns (assignment, picked_addr, channel_addrs).
    """
    n_channels = 1 + TOTAL_MUX_CHANNELS  # 13

    channel_addrs = defaultdict(set)  # ch -> set of picked addresses
    direct_addrs = set()              # mirror of channel_addrs[0]
    assignment = {}                   # comp_dir -> channel
    picked_addr = {}                  # comp_dir -> int address

    # ── Build "who else wants this address?" map ──
    addr_users = defaultdict(set)  # addr -> set of comp indices
    for i, comp in enumerate(components):
        for a in comp["usable_addresses"]:
            addr_users[a].add(i)

    # ── Phase 1: direct bus — truly-unique addresses ──
    # An address is "unique" if exactly one component lists it as usable.
    placed_indices = set()
    for i, comp in enumerate(components):
        usable = comp["usable_addresses"]
        if not usable:
            continue
        # Find addresses where this is the ONLY user
        unique_addrs = [a for a in usable if len(addr_users[a]) == 1]
        if unique_addrs:
            # Pick the first unique address
            addr = unique_addrs[0]
            channel_addrs[0].add(addr)
            direct_addrs.add(addr)
            assignment[comp["dir"]] = 0
            picked_addr[comp["dir"]] = addr
            placed_indices.add(i)

    # ── Phase 2: mux channels for everything else ──
    remaining = [i for i in range(len(components)) if i not in placed_indices]
    # Sort: fewest usable addresses first (most constrained)
    remaining.sort(key=lambda i: (len(components[i]["usable_addresses"]),
                                   components[i]["dir"]))

    for i in remaining:
        comp = components[i]
        usable = comp["usable_addresses"]
        placed = False

        if not usable:
            assignment[comp["dir"]] = -1
            picked_addr[comp["dir"]] = None
            continue

        # Try mux channels 1..12, then direct bus (0) as last resort
        channel_order = list(range(1, n_channels)) + [0]
        for ch in channel_order:
            for addr in usable:
                # Already taken on this channel?
                if addr in channel_addrs[ch]:
                    continue
                # Mux channel: can't clash with direct bus
                if ch > 0 and addr in direct_addrs:
                    continue
                # Direct bus: can't clash with any mux channel
                if ch == 0:
                    if any(addr in channel_addrs[mch]
                           for mch in range(1, n_channels)):
                        continue

                channel_addrs[ch].add(addr)
                if ch == 0:
                    direct_addrs.add(addr)
                assignment[comp["dir"]] = ch
                picked_addr[comp["dir"]] = addr
                placed = True
                break
            if placed:
                break

        if not placed:
            assignment[comp["dir"]] = -1
            picked_addr[comp["dir"]] = None

    return assignment, picked_addr, channel_addrs


def channel_label(ch):
    """Human-readable label for a channel number."""
    if ch == 0:
        return "Direct Bus (No Mux)"
    elif ch <= 8:
        m = MUXES[0]
        return f"{m['type']} (0x{m['address']:02X}) Ch{ch - 1}"
    else:
        m = MUXES[1]
        return f"{m['type']} (0x{m['address']:02X}) Ch{ch - 9}"


def channel_short_label(ch):
    """Short label for JSON export."""
    if ch == 0:
        return "direct"
    elif ch <= 8:
        return f"8ch_mux_ch{ch - 1}"
    else:
        return f"4ch_mux_ch{ch - 9}"


def channel_fill(ch):
    """Background colour for a channel."""
    if ch == 0:
        return NOMUX_FILL
    elif ch <= 8:
        return CHANNEL_FILLS_MUX1[(ch - 1) % len(CHANNEL_FILLS_MUX1)]
    else:
        return CHANNEL_FILLS_MUX2[(ch - 9) % len(CHANNEL_FILLS_MUX2)]


def channel_header_fill(ch):
    """Header colour for a channel."""
    if ch == 0:
        return NOMUX_HEADER_FILL
    elif ch <= 8:
        return MUX_HEADER_FILL
    else:
        return MUX2_HEADER_FILL


# ── Spreadsheet generation ───────────────────────────────────────────────────

def write_sheet1(ws, components, addr_map, conflicts):
    """Component Matrix + Addresses sheet."""
    ws.title = "Component Matrix"

    all_addrs = sorted(set(a for c in components for a in c["all_addresses"]))

    headers = [
        "Component", "Display Name", "Vendor", "Published",
        "I2C Addresses", "# Addrs", "Usable (excl mux)",
        "Has Jumper", "Jumper Info", "Guide URL",
        "Conflicts With", "# Conflicts", "Sensor Types",
    ]
    for a in all_addrs:
        h = f"0x{a:02X}"
        if a in MUX_RESERVED:
            h += " MUX"
        headers.append(h)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, comp in enumerate(components, 2):
        conflict_list = conflicts.get(comp["dir"], [])
        is_conflicted = len(conflict_list) > 0

        has_j = comp["has_jumper"]
        jumper_str = "Yes" if has_j else ("No" if has_j is False else "?")
        vals = [
            comp["dir"],
            comp["displayName"],
            comp["vendor"],
            "Yes" if comp["published"] else "NO",
            ", ".join(f"0x{a:02X}" for a in comp["all_addresses"]),
            len(comp["all_addresses"]),
            ", ".join(f"0x{a:02X}" for a in comp["usable_addresses"]),
            jumper_str,
            comp["jumper_info"],
            comp["guide_url"],
            ", ".join(conflict_list) if conflict_list else "None",
            len(conflict_list),
            ", ".join(comp["sensors"]) if comp["sensors"] else "",
        ]

        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 4 and not comp["published"]:
                # Make the "NO" Published cell pop out of the faint row tint.
                cell.fill = UNPUBLISHED_CELL_FILL
                cell.font = Font(bold=True, color="7F4F00")
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif not comp["published"]:
                cell.fill = UNPUBLISHED_FILL
            elif is_conflicted and col in (5, 11):
                cell.fill = CONFLICT_FILL
            elif not is_conflicted and col == 5:
                cell.fill = UNIQUE_FILL

        # Address heatmap columns
        addr_col_start = len(vals) + 1
        for ai, a in enumerate(all_addrs):
            col = addr_col_start + ai
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if a in comp["all_addresses"]:
                n_sharing = len(addr_map[a])
                cell.value = n_sharing
                if a in MUX_RESERVED:
                    cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif n_sharing > 1:
                    cell.fill = CONFLICT_FILL
                    cell.font = Font(bold=True, color="9C0006")
                else:
                    cell.fill = UNIQUE_FILL
                    cell.font = Font(color="006100")

    col_widths = {1: 18, 2: 28, 3: 26, 4: 10, 5: 40, 6: 8, 7: 36,
                  8: 10, 9: 50, 10: 40, 11: 50, 12: 10, 13: 30}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for ai in range(len(all_addrs)):
        ws.column_dimensions[get_column_letter(len(col_widths) + 1 + ai)].width = 6.5

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(components) + 1}"


def _jumper_setting(comp, addr):
    """Return (short_setting, full_info) for reaching a specific address.

    short_setting: e.g. "A0:1 A1:0", "ADDR:closed", "SDO:GND"
    full_info: the full jumper_info text
    """
    import re

    if addr is None:
        return ("", "")

    addr_hex = f"0x{addr:02X}"
    addr_hex_upper = addr_hex.upper()
    jumper_addrs = comp.get("jumper_addrs", {})
    full_info = comp.get("jumper_info", "")
    has_jumper = comp.get("has_jumper")

    # No jumper info at all
    if not full_info:
        return ("mux isolation", "No jumper info available")

    # has_jumper=False but might still have pin-selectable addresses (e.g. SHT3x ADR pin)
    # So still try the address lookup below before giving up

    # Look up the per-address description from the JSON cache
    desc = ""
    for k, v in jumper_addrs.items():
        if k.upper() == addr_hex_upper:
            desc = v
            break
    # Also try if addr falls within a range key like "0x28-0x2F"
    if not desc:
        for k, v in jumper_addrs.items():
            if "-" in k:
                try:
                    lo, hi = k.split("-")
                    if int(lo, 16) <= addr <= int(hi, 16):
                        desc = v
                        break
                except ValueError:
                    pass

    if not desc:
        if has_jumper is False:
            return ("mux isolation", full_info or "Fixed address — no jumper")
        # Address not in cache — truncate full info
        return (full_info[:50] + "...", full_info)

    # ── Convert description to concise pad notation ──
    d = desc.lower()

    if "default" in d:
        return ("default", full_info)

    # SDO/ADDR single-jumper patterns
    if "sdo" in d:
        if "gnd" in d:
            return ("SDO:GND", full_info)
        elif "bridged" in d or "closed" in d:
            return ("SDO:GND", full_info)
        elif "high" in d or "vdd" in d:
            return ("SDO:VDD", full_info)
        return ("SDO:closed", full_info)

    if "addr" in d and ("closed" in d or "bridged" in d):
        return ("ADDR:closed", full_info)
    if "addr" in d and ("vdd" in d or "vin" in d or "high" in d):
        return ("ADDR:VDD", full_info)

    if "sa0" in d:
        return ("SA0:closed" if "closed" in d else "SA0:open", full_info)

    if "adr" in d:
        if "vin" in d or "vdd" in d or "high" in d:
            return ("ADR:VDD", full_info)
        return ("ADR:GND", full_info)

    # Multi-pad patterns: A0, A1, A2, AD0, AD1
    pads = re.findall(r'\b(A[D]?\d)\b', desc, re.IGNORECASE)
    if pads:
        states = []
        for p in pads:
            states.append(f"{p.upper()}:1")
        return (" ".join(states), full_info)

    # Named jumpers like "43k jumper closed"
    jumper_match = re.search(r'(\w+)\s+jumper\s+(closed|open|bridged)', desc, re.IGNORECASE)
    if jumper_match:
        return (f"{jumper_match.group(1)}:{jumper_match.group(2)}", full_info)

    # Fallback: use the description as-is but trimmed
    return (desc[:40], full_info)


def write_sheet2(ws, components, assignment, picked_addr, channel_addrs):
    """HIL Mux Layout sheet."""
    ws.title = "HIL Mux Layout"

    n_channels = 1 + TOTAL_MUX_CHANNELS

    # Organise by channel
    channels = defaultdict(list)
    unplaceable = []
    for comp in components:
        ch = assignment.get(comp["dir"], -1)
        if ch < 0:
            unplaceable.append(comp)
        else:
            channels[ch].append(comp)

    for ch in channels:
        channels[ch].sort(key=lambda c: (picked_addr.get(c["dir"], 0) or 0, c["dir"]))

    # ── Summary section ──
    row = 1
    ws.cell(row=row, column=1, value="HIL I2C Mux Testing Layout").font = Font(
        name="Calibri", bold=True, size=14
    )
    row += 1
    ws.cell(row=row, column=1, value=f"Total components: {len(components)}")
    row += 1
    placed = sum(len(v) for v in channels.values())
    ws.cell(row=row, column=1, value=f"Placed: {placed}  |  Unplaceable: {len(unplaceable)}")
    row += 1

    for m in MUXES:
        ws.cell(row=row, column=1,
                value=f"{m['label']}: {m['type']} @ 0x{m['address']:02X} — {m['channels']} channels")
        row += 1

    ws.cell(row=row, column=1,
            value=f"Reserved addresses (mux chips): {', '.join(f'0x{a:02X}' for a in sorted(MUX_RESERVED))}")
    row += 1

    # Channel utilisation summary
    row += 1
    ws.cell(row=row, column=1, value="Channel Summary").font = Font(bold=True, size=11)
    row += 1
    sum_headers = ["Channel", "Label", "Components", "Addresses Used"]
    for hi, h in enumerate(sum_headers):
        cell = ws.cell(row=row, column=hi + 1, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = THIN_BORDER
    row += 1
    for ch in range(n_channels):
        comps_in = channels.get(ch, [])
        addrs_in = channel_addrs.get(ch, set())
        vals = [ch, channel_label(ch), len(comps_in),
                ", ".join(f"0x{a:02X}" for a in sorted(addrs_in))]
        for hi, v in enumerate(vals):
            cell = ws.cell(row=row, column=hi + 1, value=v)
            cell.border = THIN_BORDER
            cell.fill = channel_fill(ch)
        row += 1

    # Unplaceable warnings
    if unplaceable:
        row += 1
        cell = ws.cell(row=row, column=1,
                       value="UNPLACEABLE — all addresses blocked by mux reserved addresses:")
        cell.font = Font(bold=True, color="CC0000")
        row += 1
        for comp in unplaceable:
            ws.cell(row=row, column=1,
                    value=f"  {comp['dir']} ({comp['displayName']}) — "
                          f"addresses: {', '.join(f'0x{a:02X}' for a in comp['all_addresses'])}")
            row += 1

    row += 2

    # ── Channel blocks side-by-side ──
    BLOCK_WIDTH = 7
    block_headers = ["#", "Component", "Display Name", "Assigned Addr", "Jumper Setting", "All Addrs", "Vendor"]
    layout_start_row = row

    active_channels = sorted(ch for ch in range(n_channels) if channels.get(ch))

    for block_idx, ch in enumerate(active_channels):
        col_offset = block_idx * BLOCK_WIDTH
        comps_in = channels.get(ch, [])

        # Header
        r = layout_start_row
        lbl = channel_label(ch)
        for hi in range(BLOCK_WIDTH):
            c = col_offset + hi + 1
            cell = ws.cell(row=r, column=c, value=lbl if hi == 0 else "")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = channel_header_fill(ch)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=r, start_column=col_offset + 1,
                       end_row=r, end_column=col_offset + BLOCK_WIDTH)

        # Info
        r += 1
        info = f"{len(comps_in)} components"
        for hi in range(BLOCK_WIDTH):
            c = col_offset + hi + 1
            cell = ws.cell(row=r, column=c, value=info if hi == 0 else "")
            cell.font = Font(italic=True, size=9)
            cell.fill = channel_fill(ch)
            cell.border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=col_offset + 1,
                       end_row=r, end_column=col_offset + BLOCK_WIDTH)

        # Column headers
        r += 1
        for hi, hdr in enumerate(block_headers):
            c = col_offset + hi + 1
            cell = ws.cell(row=r, column=c, value=hdr)
            cell.font = Font(bold=True, size=10)
            cell.fill = channel_fill(ch)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # Component rows
        for ci, comp in enumerate(comps_in):
            r += 1
            pa = picked_addr.get(comp["dir"])
            default_addr = comp["all_addresses"][0] if comp["all_addresses"] else None
            is_non_default = pa is not None and pa != default_addr
            short_setting, _ = _jumper_setting(comp, pa) if is_non_default else ("", "")
            vals = [
                ci + 1,
                comp["dir"],
                comp["displayName"],
                f"0x{pa:02X}" if pa is not None else "?",
                short_setting,
                ", ".join(f"0x{a:02X}" for a in comp["all_addresses"]),
                comp["vendor"],
            ]
            for hi, v in enumerate(vals):
                c = col_offset + hi + 1
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if not comp["published"]:
                    cell.fill = UNPUBLISHED_FILL
                elif is_non_default and hi == 3:  # Assigned Addr column
                    cell.fill = NON_DEFAULT_FILL
                    cell.font = NON_DEFAULT_FONT
                elif is_non_default and hi == 4:  # Jumper Setting column
                    cell.fill = NON_DEFAULT_FILL
                    cell.font = NON_DEFAULT_FONT

        # Column widths
        ws.column_dimensions[get_column_letter(col_offset + 1)].width = 4
        ws.column_dimensions[get_column_letter(col_offset + 2)].width = 18
        ws.column_dimensions[get_column_letter(col_offset + 3)].width = 24
        ws.column_dimensions[get_column_letter(col_offset + 4)].width = 12
        ws.column_dimensions[get_column_letter(col_offset + 5)].width = 30
        ws.column_dimensions[get_column_letter(col_offset + 6)].width = 30
        ws.column_dimensions[get_column_letter(col_offset + 7)].width = 22

    # ── Linear ordered list for JSON export ──
    row2 = layout_start_row
    linear_col = len(active_channels) * BLOCK_WIDTH + 2

    ws.cell(row=row2, column=linear_col,
            value="Ordered Layout (for JSON / pytest export)").font = Font(bold=True, size=12)
    row2 += 1

    linear_headers = ["Order", "Channel#", "Channel Label", "Component", "Display Name",
                       "Assigned Address", "Default Address", "Jumper Setting",
                       "All Addresses", "Has Jumper", "Jumper Details", "Guide URL",
                       "Vendor", "Published", "Non-Default?"]
    for hi, hdr in enumerate(linear_headers):
        cell = ws.cell(row=row2, column=linear_col + hi, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row2 += 1

    order_num = 1
    for ch in range(n_channels):
        for comp in channels.get(ch, []):
            pa = picked_addr.get(comp["dir"])
            default_addr = comp["all_addresses"][0] if comp["all_addresses"] else None
            is_non_default = pa is not None and pa != default_addr
            has_j = comp["has_jumper"]
            jumper_str = "Yes" if has_j else ("No" if has_j is False else "?")
            short_setting, full_info = _jumper_setting(comp, pa) if is_non_default else ("", "")
            vals = [
                order_num, ch, channel_short_label(ch),
                comp["dir"], comp["displayName"],
                f"0x{pa:02X}" if pa is not None else "?",
                f"0x{default_addr:02X}" if default_addr is not None else "?",
                short_setting,
                ", ".join(f"0x{a:02X}" for a in comp["all_addresses"]),
                jumper_str,
                full_info,
                comp["guide_url"],
                comp["vendor"],
                "yes" if comp["published"] else "no",
                "NON-DEFAULT" if is_non_default else "",
            ]
            for hi, v in enumerate(vals):
                cell = ws.cell(row=row2, column=linear_col + hi, value=v)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = channel_fill(ch)
                # Published (13): highlight unpublished "no" cells
                if hi == 13 and not comp["published"]:
                    cell.fill = UNPUBLISHED_CELL_FILL
                    cell.font = Font(bold=True, color="7F4F00")
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                # Jumper Setting (7) and Non-Default (14)
                elif is_non_default and hi in (5, 7, 14):
                    cell.fill = NON_DEFAULT_FILL
                    cell.font = NON_DEFAULT_FONT
            order_num += 1
            row2 += 1

    widths = [6, 9, 16, 18, 28, 14, 14, 16, 36, 10, 50, 40, 26, 9, 14]
    for wi, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(linear_col + wi)].width = w

    ws.freeze_panes = "A1"


def write_sheet3(ws, components, addr_map):
    """Address conflict summary sheet."""
    ws.title = "Address Conflicts"

    all_addrs = sorted(set(a for c in components for a in c["all_addresses"]))

    headers = ["Address", "Mux?", "# Components", "Components"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    row = 2
    for a in all_addrs:
        comps_at = addr_map[a]
        n = len(comps_at)
        is_mux = a in MUX_RESERVED
        ws.cell(row=row, column=1, value=f"0x{a:02X}").border = THIN_BORDER
        mux_cell = ws.cell(row=row, column=2, value="MUX" if is_mux else "")
        mux_cell.border = THIN_BORDER
        cell_n = ws.cell(row=row, column=3, value=n)
        cell_n.border = THIN_BORDER
        cell_comps = ws.cell(
            row=row, column=4,
            value=", ".join(f"{c['dir']} ({c['displayName']})" for c in comps_at),
        )
        cell_comps.border = THIN_BORDER
        cell_comps.alignment = Alignment(wrap_text=True, vertical="top")

        if is_mux:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = PatternFill(
                    start_color="BF8F00", end_color="BF8F00", fill_type="solid")
            mux_cell.font = Font(bold=True, color="FFFFFF")
        elif n > 1:
            ws.cell(row=row, column=1).fill = CONFLICT_FILL
            cell_n.fill = CONFLICT_FILL
            cell_n.font = Font(bold=True, color="9C0006")
        else:
            ws.cell(row=row, column=1).fill = UNIQUE_FILL
        row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 120
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{row - 1}"


# ── Phenomena mapping ──────────────────────────────────────────────────────
# Map sensor types to physical phenomena and suggested test fixtures.
SENSOR_TO_PHENOMENON = {
    "ambient-temp":            "Temperature",
    "ambient-temp-fahrenheit": "Temperature",
    "object-temp":             "Temperature (IR)",
    "object-temp-fahrenheit":  "Temperature (IR)",
    "humidity":                "Humidity",
    "pressure":                "Pressure",
    "altitude":                "Pressure",        # derived from pressure
    "light":                   "Light",
    "proximity":               "Proximity / ToF",
    "gas-resistance":          "Gas / Air Quality",
    "raw":                     "Gas / Air Quality",
    "tvoc":                    "Gas / Air Quality",
    "eco2":                    "Gas / Air Quality",
    "co2":                     "Gas / Air Quality",
    "voc-index":               "Gas / Air Quality",
    "nox-index":               "Gas / Air Quality",
    "pm10-std":                "Particulate Matter",
    "pm25-std":                "Particulate Matter",
    "pm100-std":               "Particulate Matter",
    "current":                 "Current / Voltage",
    "voltage":                 "Current / Voltage",
    "unitless-percent":        "Other (Percent)",
}

PHENOMENON_FIXTURES = {
    "Temperature":        "Heat source / Peltier module for warming and cooling; thermally isolated enclosure",
    "Temperature (IR)":   "IR heat lamp or warm object at known distance; background target for baseline",
    "Humidity":           "Sealed enclosure with wet sponge / desiccant; or ultrasonic humidifier",
    "Pressure":           "Sealed chamber with hand pump or syringe for pressure changes",
    "Light":              "Dimmable LED or lamp with known lux levels; light-tight enclosure for dark baseline",
    "Proximity / ToF":    "Servo-driven target at known distances; flat reflective surface",
    "Gas / Air Quality":  "Sealed chamber with known VOC source (e.g. isopropyl alcohol swab); clean air baseline",
    "Particulate Matter":  "Sealed chamber with smoke/dust source; HEPA-filtered clean air baseline",
    "Current / Voltage":  "Programmable power supply or known resistive load; DAC for voltage reference",
    "Other (Percent)":    "Depends on specific sensor (e.g. soil moisture probe in wet/dry soil)",
}

PHENOMENON_FILLS = {
    "Temperature":        PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "Temperature (IR)":   PatternFill(start_color="FFD9CC", end_color="FFD9CC", fill_type="solid"),
    "Humidity":           PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
    "Pressure":           PatternFill(start_color="E5CCFF", end_color="E5CCFF", fill_type="solid"),
    "Light":              PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
    "Proximity / ToF":    PatternFill(start_color="CCF2FF", end_color="CCF2FF", fill_type="solid"),
    "Gas / Air Quality":  PatternFill(start_color="D9FFD9", end_color="D9FFD9", fill_type="solid"),
    "Particulate Matter":  PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"),
    "Current / Voltage":  PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid"),
    "Other (Percent)":    PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
}


def _classify_component(comp):
    """Return set of phenomena a component measures."""
    phenomena = set()
    for s in comp.get("sensors", []):
        p = SENSOR_TO_PHENOMENON.get(s)
        if p:
            phenomena.add(p)
    return phenomena


def write_sheet4(ws, components, assignment, picked_addr):
    """Test Fixtures — components grouped by measured phenomena."""
    ws.title = "Test Fixtures"

    # Build phenomenon -> list of components
    phenom_comps = defaultdict(list)
    for comp in components:
        for p in _classify_component(comp):
            phenom_comps[p].append(comp)

    # Sort phenomena by component count descending
    sorted_phenomena = sorted(phenom_comps.keys(),
                              key=lambda p: (-len(phenom_comps[p]), p))

    # ── Summary table ──
    row = 1
    ws.cell(row=row, column=1,
            value="Test Fixtures by Measured Phenomenon").font = Font(
                name="Calibri", bold=True, size=14)
    row += 2

    summary_headers = ["Phenomenon", "# Components", "Suggested Test Fixture"]
    for hi, h in enumerate(summary_headers):
        cell = ws.cell(row=row, column=hi + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row += 1

    for p in sorted_phenomena:
        fill = PHENOMENON_FILLS.get(p, PatternFill())
        vals = [p, len(phenom_comps[p]), PHENOMENON_FIXTURES.get(p, "")]
        for hi, v in enumerate(vals):
            cell = ws.cell(row=row, column=hi + 1, value=v)
            cell.border = THIN_BORDER
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Identify multi-phenomenon components
    multi = [(comp, _classify_component(comp))
             for comp in components
             if len(_classify_component(comp)) > 1]

    if multi:
        row += 1
        ws.cell(row=row, column=1,
                value="Multi-Phenomenon Sensors (need multiple fixtures)").font = Font(
                    bold=True, size=12)
        row += 1
        multi_headers = ["Component", "Display Name", "Phenomena", "Channel", "Address"]
        for hi, h in enumerate(multi_headers):
            cell = ws.cell(row=row, column=hi + 1, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        row += 1
        for comp, phens in sorted(multi, key=lambda x: (-len(x[1]), x[0]["dir"])):
            ch = assignment.get(comp["dir"], -1)
            pa = picked_addr.get(comp["dir"])
            vals = [
                comp["dir"],
                comp["displayName"],
                ", ".join(sorted(phens)),
                channel_label(ch) if ch >= 0 else "unplaced",
                f"0x{pa:02X}" if pa is not None else "?",
            ]
            for hi, v in enumerate(vals):
                cell = ws.cell(row=row, column=hi + 1, value=v)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    row += 2

    # ── Per-phenomenon detail blocks ──
    ws.cell(row=row, column=1,
            value="Components by Phenomenon").font = Font(bold=True, size=14)
    row += 1

    detail_headers = ["#", "Component", "Display Name", "Sensor Types",
                      "Channel", "Address", "Jumper Setting"]

    for p in sorted_phenomena:
        row += 1
        fill = PHENOMENON_FILLS.get(p, PatternFill())
        header_fill = PatternFill(
            start_color=fill.start_color.rgb[2:] if fill.start_color and fill.start_color.rgb else "333333",
            end_color=fill.start_color.rgb[2:] if fill.start_color and fill.start_color.rgb else "333333",
            fill_type="solid")

        # Phenomenon header
        fixture_text = PHENOMENON_FIXTURES.get(p, "")
        cell = ws.cell(row=row, column=1,
                       value=f"{p}  ({len(phenom_comps[p])} components)")
        cell.font = Font(bold=True, size=12)
        cell.fill = fill
        for ci in range(len(detail_headers)):
            ws.cell(row=row, column=ci + 1).fill = fill
        row += 1

        # Fixture note
        cell = ws.cell(row=row, column=1, value=f"Fixture: {fixture_text}")
        cell.font = Font(italic=True, size=10, color="555555")
        row += 1

        # Column headers
        for hi, h in enumerate(detail_headers):
            cell = ws.cell(row=row, column=hi + 1, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Sort by channel then name
        comps_sorted = sorted(phenom_comps[p],
                              key=lambda c: (assignment.get(c["dir"], 99), c["dir"]))
        for ci, comp in enumerate(comps_sorted):
            ch = assignment.get(comp["dir"], -1)
            pa = picked_addr.get(comp["dir"])
            default_addr = comp["all_addresses"][0] if comp["all_addresses"] else None
            is_non_default = pa is not None and pa != default_addr
            short_setting, _ = _jumper_setting(comp, pa) if is_non_default else ("", "")
            vals = [
                ci + 1,
                comp["dir"],
                comp["displayName"],
                ", ".join(comp["sensors"]),
                channel_label(ch) if ch >= 0 else "unplaced",
                f"0x{pa:02X}" if pa is not None else "?",
                short_setting,
            ]
            for hi, v in enumerate(vals):
                cell = ws.cell(row=row, column=hi + 1, value=v)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if is_non_default and hi in (5, 6):
                    cell.fill = NON_DEFAULT_FILL
                    cell.font = NON_DEFAULT_FONT
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 20
    ws.freeze_panes = "A1"


def main():
    parser = argparse.ArgumentParser(
        description="Generate the HIL I2C testing spreadsheet from Wippersnapper_Components."
    )
    parser.add_argument(
        "--dual-mux", action="store_true",
        help="Also use the 2nd mux (TCA9544A @ 0x71, 4ch). "
             "Default: single TCA9548A @ 0x77 only (0x71 left as a usable address).",
    )
    parser.add_argument(
        "-o", "--output", default=None,
        help="Output .xlsx path (default: hil_i2c_components.xlsx next to this script).",
    )
    args = parser.parse_args()
    configure_muxes(args.dual_mux)

    base_dir = Path(__file__).parent
    components = load_components(base_dir)

    # Check for missing jumper info and auto-fetch if possible
    jumper_db = load_jumper_info(base_dir)
    jumper_db = check_and_fetch_missing_jumper_info(base_dir, components, jumper_db)
    # Reload components to pick up any newly fetched info
    if any(c["dir"] not in jumper_db or c.get("has_jumper") is None for c in components):
        components = load_components(base_dir)

    addr_map = build_address_map(components)
    conflicts = find_conflicts(components, addr_map)

    print(f"Loaded {len(components)} I2C components")
    print(f"Unique I2C addresses: {len(set(a for c in components for a in c['all_addresses']))}")
    print(f"Mux reserved addresses: {', '.join(f'0x{a:02X}' for a in sorted(MUX_RESERVED))}")
    print()

    for m in MUXES:
        print(f"  {m['label']}: {m['type']} @ 0x{m['address']:02X} ({m['channels']} channels)")
    print(f"  Total mux channels: {TOTAL_MUX_CHANNELS}")
    print()

    assignment, picked_addr, channel_addrs = assign_channels(components)

    n_channels = 1 + TOTAL_MUX_CHANNELS
    placed = sum(1 for v in assignment.values() if v >= 0)
    unplaced = sum(1 for v in assignment.values() if v < 0)
    print(f"Placed: {placed}  |  Unplaceable: {unplaced}")

    for ch in range(n_channels):
        comps_in = [c for c in components if assignment.get(c["dir"]) == ch]
        if not comps_in:
            continue
        addrs_in = channel_addrs.get(ch, set())
        print(f"  {channel_label(ch)}: {len(comps_in)} components, "
              f"addrs: {', '.join(f'0x{a:02X}' for a in sorted(addrs_in))}")

    unplaced_comps = [c for c in components if assignment.get(c["dir"], -1) < 0]
    if unplaced_comps:
        print("\nUNPLACEABLE:")
        for c in unplaced_comps:
            print(f"  {c['dir']} — all addresses blocked by mux: "
                  f"{', '.join(f'0x{a:02X}' for a in c['all_addresses'])}")

    # Verify no conflicts
    errors = 0
    direct_addrs = channel_addrs.get(0, set())
    for ch in range(n_channels):
        seen = {}
        comps_in = [c for c in components if assignment.get(c["dir"]) == ch]
        for comp in comps_in:
            pa = picked_addr[comp["dir"]]
            if pa in seen:
                print(f"BUG: channel {ch} addr 0x{pa:02X} used by {seen[pa]} AND {comp['dir']}")
                errors += 1
            seen[pa] = comp["dir"]
            if ch > 0 and pa in direct_addrs:
                print(f"BUG: mux ch{ch} {comp['dir']} @ 0x{pa:02X} conflicts with direct bus")
                errors += 1
    if errors == 0:
        print("\nVERIFIED: zero address conflicts across all channels")
    else:
        print(f"\n{errors} BUGS FOUND")

    # Create workbook
    wb = Workbook()
    ws1 = wb.active
    write_sheet1(ws1, components, addr_map, conflicts)

    ws2 = wb.create_sheet()
    write_sheet2(ws2, components, assignment, picked_addr, channel_addrs)

    ws3 = wb.create_sheet()
    write_sheet3(ws3, components, addr_map)

    ws4 = wb.create_sheet()
    write_sheet4(ws4, components, assignment, picked_addr)

    out_path = Path(args.output) if args.output else base_dir / "hil_i2c_components.xlsx"
    wb.save(out_path)
    print(f"\nSpreadsheet saved to: {out_path}")


if __name__ == "__main__":
    main()
